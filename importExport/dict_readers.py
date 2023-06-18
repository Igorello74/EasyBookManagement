"""Readers of different data formats, such as xlsx, csv, json, etc.

These readers should provide DictReader interface, which is an iterator,
yielding dicts with data from the current row. An example: csv.DictReader
"""

import csv
import pathlib

from openpyxl import load_workbook

from utils import ObjectFactory


class BadFileError(Exception):
    ...


class XlsxDictReader:
    def __init__(self, f, fieldnames=None, *args, **kwds):
        self._fieldnames = fieldnames   # list of keys for the dict
        self.reader = load_workbook(
            f, read_only=True, data_only=True, *args, **kwds
        ).active.iter_rows(values_only=True)
        self.line_num = 0

    def __iter__(self):
        return self

    @property
    def fieldnames(self):
        if self._fieldnames is None:
            try:
                self._fieldnames = [str(i).lower() for i in next(self.reader)]
            except StopIteration:
                pass
        self.line_num = 1
        return self._fieldnames

    @fieldnames.setter
    def fieldnames(self, value):
        self._fieldnames = value

    def __next__(self):
        if self.line_num == 0:
            # Used only for its side effect.
            self.fieldnames
        row = next(self.reader)

        # unlike the basic reader, we prefer not to return blanks,
        # because we will typically wind up with a dict full of None
        # values
        while row == []:
            row = next(self.reader)
        d = dict(zip(self.fieldnames, row))
        return d


class CsvDictReader(csv.DictReader):

    def __init__(self, f, encoding="utf-8", *args, **kwargs):
        if encoding:
            f = self._decode_file(f, encoding)
        super().__init__(f, *args, **kwargs)

    @staticmethod
    def _decode_file(file, encoding='utf-8'):
        for i in file:
            yield i.decode(encoding, "replace")

    @property
    def fieldnames(self):
        if self._fieldnames is None:
            try:
                self._fieldnames = [str(i).lower() for i in next(self.reader)]
            except StopIteration:
                pass
        self.line_num = self.reader.line_num
        return self._fieldnames

    @fieldnames.setter
    def fieldnames(self, value):
        self._fieldnames = value


class DictReaderFactory(ObjectFactory):
    def get(self, file, **kwargs):
        file_extension = pathlib.Path(file.name).suffix
        try:
            return super().get(key=file_extension, f=file, **kwargs)
        except ValueError:
            raise BadFileError(
                f"We can't read that format yet: {file_extension}")
        except Exception:
            raise BadFileError("Invalid file")


factory = DictReaderFactory()
factory.register(".csv", CsvDictReader)
factory.register(".xlsx", XlsxDictReader)
