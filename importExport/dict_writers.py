"""Writers of different data formats, such as xlsx, csv, json, etc.

These writers should provide DictWriter interface, which is an iterator,
yielding dicts with data from the current row. An example: csv.DictWriter
"""

import csv

import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

from EasyBookManagement import ObjectFactory


class DictWriterFactory(ObjectFactory):
    ...


class XlsxDictWriter(csv.DictWriter):
    def __init__(self, f, fieldnames, restval="", *args, **kwargs):
        self.filename = f.name
        self.extrasaction = "ignore"
        self.fieldnames = fieldnames    # list of keys for the dict
        self.restval = restval          # for writing short dicts

        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.wrap_multiline_cells = kwargs.get("wrap_multiline_cells", False)
        self.wrapped_line_width = kwargs.get("wrapped_line_width", 150)

    def writerow(self, rowdict):
        return self.ws.append(self._dict_to_list(rowdict))

    def writerows(self, rowdicts):
        return (self.writerow(self._dict_to_list(i)) for i in rowdicts)

    def save(self):
        if self.wrap_multiline_cells:
            cols_with_multilines = set()

            for row in self.ws.iter_rows():
                for col_ind, cell in enumerate(row, 1):
                    if "\n" in str(cell.value):
                        cell.alignment = openpyxl.styles.Alignment(
                            wrap_text=True)
                        cols_with_multilines.add(col_ind)

            for col in cols_with_multilines:
                col_letter = get_column_letter(col)
                self.ws.column_dimensions[col_letter].width = self.wrapped_line_width

        self.wb.save(self.filename)


class CsvDictWriter(csv.DictWriter):
    def save(self):
        pass


factory = DictWriterFactory()
factory.register(".csv", CsvDictWriter)
factory.register(".xlsx", XlsxDictWriter)
