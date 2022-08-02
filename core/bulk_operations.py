"""Implement some bulk operations on models,
such as import/export from/to json, csv, xlsx etc.
"""

import csv
from zipfile import BadZipFile

from openpyxl import load_workbook
from django.db import models


class XlsxDictReader:
    def __init__(self, f, fieldnames=None, restkey=None,
                 restval=None, *args, **kwds):
        self._fieldnames = fieldnames   # list of keys for the dict
        self.restkey = restkey          # key to catch long rows
        self.restval = restval          # default value for short rows
        self.reader = load_workbook(
            f, read_only=True, data_only=True, *args, **kwds
        ).active.iter_rows(values_only=True)
        self.line_num = 0

    def __iter__(self):
        return self

    @property
    def fieldnames(self):
        if self._fieldnames is None:
            try:
                self._fieldnames = [str(i).lower() for i in next(self.reader)]
            except StopIteration:
                pass
        self.line_num = 1
        return self._fieldnames

    @fieldnames.setter
    def fieldnames(self, value):
        self._fieldnames = value

    def __next__(self):
        if self.line_num == 0:
            # Used only for its side effect.
            self.fieldnames
        row = next(self.reader)

        # unlike the basic reader, we prefer not to return blanks,
        # because we will typically wind up with a dict full of None
        # values
        while row == []:
            row = next(self.reader)
        d = dict(zip(self.fieldnames, row))
        lf = len(self.fieldnames)
        lr = len(row)
        if lf < lr:
            d[self.restkey] = row[lf:]
        elif lf > lr:
            for key in self.fieldnames[lr:]:
                d[key] = self.restval
        return d


def decode_file(file):
    for i in file:
        yield i.decode("utf-8", "replace")


class CsvDictReader(csv.DictReader):
    @property
    def fieldnames(self):
        if self._fieldnames is None:
            try:
                self._fieldnames = [str(i).lower() for i in next(self.reader)]
            except StopIteration:
                pass
        self.line_num = self.reader.line_num
        return self._fieldnames

    @fieldnames.setter
    def fieldnames(self, value):
        self._fieldnames = value


class BadFile(Exception):
    ...


class ColumnNotFoundError(LookupError):
    def __init__(self, missing_columns: list):
        self.missing_columns = missing_columns

        message = f"{', '.join(missing_columns)} not present in your sheet"
        super().__init__(message)


class BulkHandler:
    """Implement some bulk operations on models,
    such as import/export from/to json, csv, xlsx etc.
    """

    def __init__(self, model: models.Model, headers_mapping: dict,
                 pk_field="id"):
        """Create an instance of the BulkHandler class.

        model: a django model class on which operations are held;
        headers_mapping: a mapping of model fields
          to column headers (custom column names);
        pk_field: name of model's primary key (usually it is"id")
        """
        self.model = model
        self.headers_mapping = {
            k: v.lower() for k, v in headers_mapping.items()
        }
        self.pk_field = pk_field

    def create_or_update(self, file, required_fields: list = []) -> dict:
        """Create or update model instancies from an xlsx or csv file

        file: a file-like object

        If id is not provided in a row, it's considered to be a new reader,
        if id is present, this entry is updated.

        If any of required_fields is absent, a ColumnNotFoundError is raised.
        Note: headers are case insensitive.

        Returns {"created": "<num of instancies created>",
                "updated": "<num of instancies updated>"}
        """

        try:
            file_reader = XlsxDictReader(file)
        except BadZipFile:
            try:
                file_reader = CsvDictReader(decode_file(file))
            except csv.Error:
                raise BadFile("Invalid file, you must provide csv or xlsx")

        id_col = self.headers_mapping.get(self.pk_field, self.pk_field)

        # Iterate through rows and save models
        objs_to_create = []  # rows with no id provided are appended to it
        objs_to_update = []  # rows with id are appended to it

        for column in file_reader:
            obj = self.model()
            id = column.get(self.pk_field)
            if id != "" and id is not None:
                # if id is present, it's an update operation
                objs_to_update.append(obj)
                obj.id = column[id_col]
            else:
                objs_to_create.append(obj)

            for field, col_name in self.headers_mapping.items():
                # field is the standard field name from self.fields,
                # col_name is custom column name mapped to the field
                if val := column.get(col_name):
                    # if this column is present in the table
                    setattr(obj, field, val)

            absent_columns = []
            for i in required_fields:
                if not getattr(obj, i):
                    absent_columns.append(i)
            if absent_columns:
                raise ColumnNotFoundError(absent_columns)

        created = updated = 0

        if objs_to_create:
            created = len(self.model.objects.bulk_create(objs_to_create))
        if objs_to_update:
            fields = [k for k, v in self.headers_mapping.items()
                      if v in file_reader.fieldnames and k != self.pk_field]

            updated = self.model.objects.bulk_update(objs_to_update, fields)

        return {'created': created, 'updated': updated}
