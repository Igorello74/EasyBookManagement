"""Implement some bulk operations on books,
such as import/export from/to json, csv, xlsx etc.
"""

import csv
from zipfile import BadZipFile

from openpyxl import load_workbook

from .models import Reader


class XlsxDictReader:
    def __init__(self, f, fieldnames=None, restkey=None,
                 restval=None, *args, **kwds):
        self._fieldnames = fieldnames   # list of keys for the dict
        self.restkey = restkey          # key to catch long rows
        self.restval = restval          # default value for short rows
        self.reader = load_workbook(
            f, read_only=True, data_only=True, *args, **kwds
        ).active.iter_rows(values_only=True)
        self.line_num = 0

    def __iter__(self):
        return self

    @property
    def fieldnames(self):
        if self._fieldnames is None:
            try:
                self._fieldnames = [str(i).lower() for i in next(self.reader)]
            except StopIteration:
                pass
        self.line_num = 1
        return self._fieldnames

    @fieldnames.setter
    def fieldnames(self, value):
        self._fieldnames = value

    def __next__(self):
        if self.line_num == 0:
            # Used only for its side effect.
            self.fieldnames
        row = next(self.reader)

        # unlike the basic reader, we prefer not to return blanks,
        # because we will typically wind up with a dict full of None
        # values
        while row == []:
            row = next(self.reader)
        d = dict(zip(self.fieldnames, row))
        lf = len(self.fieldnames)
        lr = len(row)
        if lf < lr:
            d[self.restkey] = row[lf:]
        elif lf > lr:
            for key in self.fieldnames[lr:]:
                d[key] = self.restval
        return d


def decode_file(file):
    for i in file:
        yield i.decode("utf-8", "replace")


class CsvDictReader(csv.DictReader):
    @property
    def fieldnames(self):
        if self._fieldnames is None:
            try:
                self._fieldnames = [str(i).lower() for i in next(self.reader)]
            except StopIteration:
                pass
        self.line_num = self.reader.line_num
        return self._fieldnames

    @fieldnames.setter
    def fieldnames(self, value):
        self._fieldnames = value


class BadFile(Exception):
    ...


class ColumnNotFoundError(LookupError):
    def __init__(self, missing_columns: list):
        self.missing_columns = missing_columns

        message = f"{', '.join(missing_columns)} not present in your sheet"
        super().__init__(message)


def import_readers(file, headers: dict) -> dict:
    """Create or update readers from an xlsx or csv file

    file: a file-like object

    headers: a mapping of headers (see below)
    to columns (your custom column names)

    If id is not provided in a row, it's considered to be a new reader,
    if id is present, this entry is updated.

    If name column is absent, a ColumnNotFoundError is raised.
    If role is not provided, it's set STUDENT by default.
    Note: headers are case insensitive.

    Returns {"created": "<num of readers created>",
             "updated": "<num of readers updated>"}

    Possible headers:
        name, role, group, profile, notes,
        first_lang (ISO 639-1, like en, ru), second_lang
    """
    POSSIBLE_HEADERS = ("id", "name", "role", "group", "profile",
                        "notes", "first_lang", "second_lang")
    headers = {k: v.lower()
               for k, v in headers.items() if k in POSSIBLE_HEADERS}
    assert "name" in headers, "Name is required!"

    try:
        file_reader = XlsxDictReader(file)
    except BadZipFile:
        try:
            file_reader = CsvDictReader(decode_file(file))
        except csv.Error:
            raise BadFile("Invalid file, you must provide csv or xlsx")

    id_col = headers.get('id', 'id')

    # Iterate through rows and save models
    objs_to_create = []  # rows with no id provided are appended to it
    objs_to_update = []  # rows with id are appended to it

    for column in file_reader:
        r = Reader()
        if id_col in column:
            # if id is present, it's an update operation
            objs_to_update.append(r)
            r.id = column[id_col]
            create = False
        else:
            objs_to_create.append(r)
            create = True

        for field, col_name in headers.items():
            # field is the standard field name from POSSIBLE_HEADERS,
            # col_name is custom column name mapped to the field
            if val := column.get(col_name):
                # if this column is present in the table
                setattr(r, field, val)

        if not r.role and create:
            r.role = Reader.STUDENT
        if not r.name:
            raise ColumnNotFoundError(['name'])

    created = updated = 0

    if objs_to_create:
        created = len(Reader.objects.bulk_create(objs_to_create))
    if objs_to_update:
        fields = [k for k, v in headers.items()
                  if v in file_reader.fieldnames and k != "id"]

        updated = Reader.objects.bulk_update(objs_to_update, fields)

    return {'created': created, 'updated': updated}
